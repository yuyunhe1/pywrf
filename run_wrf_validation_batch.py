#!/usr/bin/env python3
"""Run independent daily WRF forecasts, then validate only available observations."""

import argparse
import csv
import math
import os
import re
import subprocess
import sys
import time
import warnings
from datetime import datetime, timedelta
from pathlib import Path

from utils import parse_gdex_gfs_file


warnings.filterwarnings("ignore", message="Workbook contains no default style")

OBS_DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})")
OBS_COMPACT_DATE_PATTERN = re.compile(r"(\d{8})")


def parse_day(value):
    return datetime.strptime(value, "%Y-%m-%d")


def date_range(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def observation_date_and_format(path):
    if path.parent.name == "202511~202602" or path.name.startswith("WindData_product_"):
        match = OBS_COMPACT_DATE_PATTERN.search(path.name)
        return (datetime.strptime(match.group(1), "%Y%m%d"), "old") if match else (None, None)
    if re.fullmatch(r"\d{4}-\d{2}", path.parent.name):
        match = OBS_DATE_PATTERN.search(path.name)
        return (parse_day(match.group(1)), "new") if match else (None, None)
    return None, None


def old_file_has_core_observations(
    path, min_height=30.0, max_height=120.0, min_valid_fraction=0.5
):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("检查old格式低空实测需要openpyxl：conda install -c conda-forge openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        # These exported workbooks report A1 as the used range; reset so
        # read-only iteration sees all wind-height columns.
        sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, ())
        speed_columns = []
        for index, value in enumerate(header):
            try:
                height = float(value)
            except (TypeError, ValueError):
                continue
            if min_height <= height <= max_height and index > 0 and (index - 1) % 3 == 0:
                speed_columns.append(index)

        valid_count = 0
        possible_count = 0
        for row in rows:
            for index in speed_columns:
                possible_count += 1
                if index >= len(row):
                    continue
                try:
                    value = float(row[index])
                except (TypeError, ValueError):
                    continue
                if math.isfinite(value) and value != -999:
                    valid_count += 1
        return (
            valid_count > 0
            and possible_count > 0
            and valid_count / possible_count >= min_valid_fraction
        )
    finally:
        workbook.close()


def observation_days(
    base_dir, start, end, obs_format, check_core_content=True, min_valid_fraction=0.5
):
    file_days = set()
    core_days = set()
    for path in Path(base_dir).rglob("*.xls*"):
        day, file_format = observation_date_and_format(path)
        if day is None or not (start <= day <= end):
            continue
        expected_format = "old" if obs_format == "auto" and day.month in {11, 12, 1, 2} else obs_format
        if expected_format != "auto" and file_format != expected_format:
            continue
        file_days.add(day)
        if (
            not check_core_content
            or file_format == "new"
            or old_file_has_core_observations(path, min_valid_fraction=min_valid_fraction)
        ):
            core_days.add(day)
    return file_days, core_days


def build_gfs_index(gfs_dir):
    valid_times = set()
    cycle_valid_times = {}
    for path in Path(gfs_dir).rglob("gfs.0p25.*.f*.grib2"):
        parsed = parse_gdex_gfs_file(path)
        if parsed is None:
            continue
        cycle_time, _, valid_time = parsed
        valid_times.add(valid_time)
        cycle_valid_times.setdefault(cycle_time, set()).add(valid_time)
    return valid_times, cycle_valid_times


def required_times(day, forecast_hours, interval_hours):
    return [
        day + timedelta(hours=offset)
        for offset in range(0, forecast_hours + 1, interval_hours)
    ]


def cycle_for_day(day, cycle_hour):
    cycle = day.replace(hour=cycle_hour)
    return cycle if cycle <= day else cycle - timedelta(days=1)


def gfs_day_available(day, forecast_hours, interval_hours, valid_times, cycle_valid_times, cycle_hour):
    expected = required_times(day, forecast_hours, interval_hours)
    if cycle_hour is None:
        return all(value in valid_times for value in expected)
    cycle = cycle_for_day(day, cycle_hour)
    return all(value in cycle_valid_times.get(cycle, set()) for value in expected)


def evenly_sample(days, maximum):
    days = sorted(days)
    if maximum <= 0 or len(days) <= maximum:
        return days
    if maximum == 1:
        return [days[len(days) // 2]]
    indexes = [round(index * (len(days) - 1) / (maximum - 1)) for index in range(maximum)]
    return [days[index] for index in sorted(set(indexes))]


def stratified_sample(days, maximum):
    days = sorted(days)
    if maximum <= 0 or len(days) <= maximum:
        return days

    by_month = {}
    for day in days:
        by_month.setdefault((day.year, day.month), []).append(day)
    if maximum < len(by_month):
        return evenly_sample(days, maximum)

    selected = {month_days[len(month_days) // 2] for month_days in by_month.values()}
    remaining = [day for day in days if day not in selected]
    remaining_slots = maximum - len(selected)
    if remaining_slots > 0:
        selected.update(evenly_sample(remaining, remaining_slots))
    return sorted(selected)


def output_exists(base_dir, day, domain):
    folder = Path(base_dir, "wrf_output", day.strftime("%Y-%m-%d"))
    return any(folder.glob(f"wrfout_d{domain:02d}_*"))


def write_status(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["date", "status", "elapsed_hours", "return_code", "message"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_plan(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "date",
        "observation_file_available",
        "core_observation_available",
        "gfs_complete",
        "existing_wrfout",
        "selected",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def run_command(command, cwd, env=None):
    started = time.time()
    result = subprocess.run(command, cwd=str(cwd), env=env)
    return result.returncode, (time.time() - started) / 3600.0


def main():
    parser = argparse.ArgumentParser(
        description="按天运行WRF，并仅使用存在的实测数据进行低空风场验证。"
    )
    parser.add_argument("--base-dir", default="/root/pyWRF-automation")
    parser.add_argument("--start-date", default="2025-11-01")
    parser.add_argument("--end-date", default="2026-02-28")
    parser.add_argument("--forecast-hours", type=int, default=24)
    parser.add_argument("--gfs-interval-hours", type=int, default=3)
    parser.add_argument("--gfs-cycle-hour", type=int, choices=[0, 6, 12, 18], default=None)
    parser.add_argument(
        "--selection",
        choices=["observed", "all"],
        default="observed",
        help="observed=只模拟存在有效30-120m实测的日期；all=模拟范围内所有日期",
    )
    parser.add_argument(
        "--obs-format",
        choices=["auto", "new", "old"],
        default="old",
        help="用于日期筛选和最终验证的实测格式；11月至2月建议old",
    )
    parser.add_argument(
        "--skip-observation-content-check",
        action="store_true",
        help="仅按实测文件是否存在选择日期，不检查30-120m是否有非-999数据",
    )
    parser.add_argument(
        "--min-observation-valid-fraction",
        type=float,
        default=0.5,
        help="选择日期所需的30-120m有效实测比例；默认至少50%%",
    )
    parser.add_argument(
        "--max-days",
        type=int,
        default=12,
        help="从候选日期均匀抽样的最大天数；0=全部候选日期",
    )
    parser.add_argument(
        "--max-wall-hours",
        type=float,
        default=24.0,
        help="达到该墙钟时间后不再启动新日期；0=不限制",
    )
    parser.add_argument("--num-proc", type=int, default=4)
    parser.add_argument("--domain", type=int, default=2)
    parser.add_argument("--rerun-existing", action="store_true")
    parser.add_argument("--skip-validation", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    start = parse_day(args.start_date)
    end = parse_day(args.end_date)
    if (
        end < start
        or args.forecast_hours <= 0
        or args.gfs_interval_hours <= 0
        or args.num_proc <= 0
        or args.max_days < 0
        or args.max_wall_hours < 0
        or not 0 <= args.min_observation_valid_fraction <= 1
    ):
        raise SystemExit("日期范围、forecast-hours或gfs-interval-hours无效")

    base_dir = Path(args.base_dir)
    gfs_dir = base_dir / "data" / "gdex_gfs_0p25_global"
    valid_times, cycle_valid_times = build_gfs_index(gfs_dir)
    candidates = set(date_range(start, end))
    observation_files, observed = observation_days(
        base_dir,
        start,
        end,
        args.obs_format,
        check_core_content=not args.skip_observation_content_check,
        min_valid_fraction=args.min_observation_valid_fraction,
    )
    if args.selection == "observed":
        candidates &= observed

    candidates_with_gfs = [
        day
        for day in candidates
        if gfs_day_available(
            day,
            args.forecast_hours,
            args.gfs_interval_hours,
            valid_times,
            cycle_valid_times,
            args.gfs_cycle_hour,
        )
    ]
    selected = stratified_sample(candidates_with_gfs, args.max_days)
    selected_set = set(selected)
    plan_path = base_dir / "data" / "wrf_batch_plan.csv"
    plan_rows = []
    for day in date_range(start, end):
        plan_rows.append(
            {
                "date": day.strftime("%Y-%m-%d"),
                "observation_file_available": day in observation_files,
                "core_observation_available": day in observed,
                "gfs_complete": gfs_day_available(
                    day,
                    args.forecast_hours,
                    args.gfs_interval_hours,
                    valid_times,
                    cycle_valid_times,
                    args.gfs_cycle_hour,
                ),
                "existing_wrfout": output_exists(base_dir, day, args.domain),
                "selected": day in selected_set,
            }
        )
    write_plan(plan_path, plan_rows)

    print(f"Date range: {start:%Y-%m-%d} to {end:%Y-%m-%d}")
    print(f"Observation-file days: {len(observation_files)}")
    print(f"Days with available 30-120m observations: {len(observed)}")
    print(f"Candidate days with complete GFS: {len(candidates_with_gfs)}")
    print(f"Selected days: {len(selected)}")
    for day in selected:
        print(f"  {day:%Y-%m-%d}")
    print(f"Batch plan: {plan_path}")
    if args.dry_run:
        return
    if not selected:
        raise SystemExit("没有满足实测选择和GFS完整性条件的日期；请先检查服务器数据目录")

    status_path = base_dir / "data" / "wrf_batch_status.csv"
    rows = []
    write_status(status_path, rows)
    batch_started = time.time()
    completed_elapsed = []
    for day in selected:
        wall_elapsed = (time.time() - batch_started) / 3600.0
        estimated_next = sum(completed_elapsed) / len(completed_elapsed) if completed_elapsed else 0.0
        if args.max_wall_hours > 0 and (
            wall_elapsed >= args.max_wall_hours
            or (estimated_next > 0 and wall_elapsed + estimated_next > args.max_wall_hours)
        ):
            rows.append(
                {
                    "date": day.strftime("%Y-%m-%d"),
                    "status": "not_started_wall_limit",
                    "elapsed_hours": "0",
                    "return_code": "",
                    "message": f"batch wall limit {args.max_wall_hours:g}h",
                }
            )
            write_status(status_path, rows)
            break

        if not args.rerun_existing and output_exists(base_dir, day, args.domain):
            rows.append(
                {
                    "date": day.strftime("%Y-%m-%d"),
                    "status": "skipped_existing",
                    "elapsed_hours": "0",
                    "return_code": "0",
                    "message": "wrfout already exists",
                }
            )
            write_status(status_path, rows)
            continue

        env = os.environ.copy()
        env["WRF_START_DATE"] = day.strftime("%Y-%m-%d_00:00:00")
        env["WRF_FORECAST_HOURS"] = str(args.forecast_hours)
        env["WRF_NUM_PROC"] = str(args.num_proc)
        env["REUSE_GEOGRID"] = "1"
        if args.gfs_cycle_hour is not None:
            env["GFS_CYCLE_TIME"] = cycle_for_day(day, args.gfs_cycle_hour).strftime(
                "%Y-%m-%d_%H:%M:%S"
            )
        else:
            env.pop("GFS_CYCLE_TIME", None)

        print(f"\n[BATCH] Running {day:%Y-%m-%d}")
        rows.append(
            {
                "date": day.strftime("%Y-%m-%d"),
                "status": "running",
                "elapsed_hours": "0",
                "return_code": "",
                "message": "",
            }
        )
        write_status(status_path, rows)
        return_code, elapsed = run_command([sys.executable, "main.py"], base_dir, env)
        completed_elapsed.append(elapsed)
        rows[-1] = {
            "date": day.strftime("%Y-%m-%d"),
            "status": "completed" if return_code == 0 else "failed",
            "elapsed_hours": f"{elapsed:.3f}",
            "return_code": str(return_code),
            "message": "",
        }
        write_status(status_path, rows)

    available_outputs = list((base_dir / "wrf_output").rglob(f"wrfout_d{args.domain:02d}_*"))
    if not args.skip_validation and available_outputs:
        validation_output = base_dir / "data" / "wrf_wind_validation_batch"
        command = [
            sys.executable,
            "compare_wrf_wind_observations.py",
            "--wrf-dir",
            str(base_dir / "wrf_output"),
            "--obs-dir",
            str(base_dir),
            "--obs-format",
            args.obs_format,
            "--output-dir",
            str(validation_output),
        ]
        print("\n[BATCH] Validating all available wrfout against available observations")
        rows.append(
            {
                "date": "VALIDATION",
                "status": "validation_running",
                "elapsed_hours": "0",
                "return_code": "",
                "message": "",
            }
        )
        write_status(status_path, rows)
        return_code, elapsed = run_command(command, base_dir)
        rows[-1] = {
            "date": "VALIDATION",
            "status": "validation_completed" if return_code == 0 else "validation_failed",
            "elapsed_hours": f"{elapsed:.3f}",
            "return_code": str(return_code),
            "message": "",
        }
        write_status(status_path, rows)
        print(f"[BATCH] Validation return code={return_code}, elapsed={elapsed:.3f}h")
    elif not args.skip_validation:
        print("[BATCH] Validation skipped because no wrfout files are available")

    print(f"[BATCH] Status: {status_path}")


if __name__ == "__main__":
    main()
